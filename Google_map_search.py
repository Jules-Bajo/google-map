

import googlemaps
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook

def search( maps , address ) :
    # Geocoding an address
    geocode_result = gmaps.geocode( address )

    if len( geocode_result ) > 0 :
        loc = geocode_result[0]['geometry']['location']
        print( str( loc['lat'] ) + ', ' + str( loc['lng'] ) )
        return loc

def find( rows , lat , lng ) :
    found = False 
    for row in rows :
        if lat == row[5] and lng == row[6] :
            found = True
            break

    return found

def loadFile( fileName , rows , startRow , endRow ) :
    wb = load_workbook( fileName )
    rng = wb['ALL']
    rowCount = endRow - startRow + 1 

    for i in range( rowCount ) :
        sa = 'B' + str( startRow + i )
        print( '-------------' + sa + '-------------' )
        print( rng[ sa ].value )
        loc = search( gmaps , rng[ sa ].value )
        if loc is not None :
            lat = loc[ 'lat' ]
            lng = loc[ 'lng' ]
            if not find( rows , lat , lng ) :
                row = []
                row.append( rng[ 'A' + str( 2 + i ) ].value )
                row.append( rng[ 'B' + str( 2 + i ) ].value )
                row.append( rng[ 'C' + str( 2 + i ) ].value )
                row.append( rng[ 'D' + str( 2 + i ) ].value )
                row.append( rng[ 'G' + str( 2 + i ) ].value )
                row.append( lat )
                row.append( lng )
                rows.append( row )

def toExcel( ws , rows ) :
    k = 2 
    for row in rows :
        ws[ 'A' + str( k ) ] = row[0]
        ws[ 'B' + str( k ) ] = row[1]
        ws[ 'C' + str( k ) ] = row[2]
        ws[ 'D' + str( k ) ] = row[3]
        ws[ 'E' + str( k ) ] = row[4]
        ws[ 'F' + str( k ) ] = row[5]
        ws[ 'G' + str( k ) ] = row[6]        
        k += 1

print( 'start' )

gmaps = googlemaps.Client(key='-----Load your API KEY------')

newWb = Workbook()

ws = newWb.active
ws.title = 'ALL'

ws[ 'A1' ] = 'STORE_CODE'
ws[ 'B1' ] = 'STORE ADDRESS'
ws[ 'C1' ] = 'City'
ws[ 'D1' ] = 'STATE'
ws[ 'E1' ] = 'Pin Code'
ws[ 'F1' ] = 'LATITUDE'
ws[ 'G1' ] = 'LONGITUDE'

rows = []

# Load Location Data

loadFile( 'Load your Data exil file XXXX.xlsx' , rows , 2 , 535 )

toExcel( ws , rows )

# save generated location lat. Log. to a file. 

newWb.save( filename = 'generated_location.xlsx' ) # the file that will be produce after generating logitude and latitude

print( 'done' )

